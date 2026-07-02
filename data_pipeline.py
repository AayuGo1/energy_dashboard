"""
Data Pipeline Hardening Subsystem — Ingestion Framework with Snapshot Fallbacks.
"""
import os
import re
import json
import logging
from io import BytesIO
from datetime import datetime
import requests
from openpyxl import load_workbook
import config

def get_logger() -> logging.Logger:
    logger = logging.getLogger("jfl_portal")
    if logger.handlers:
        return logger
    logger.setLevel(getattr(logging, config.LOG_LEVEL, logging.INFO))
    try:
        os.makedirs(os.path.dirname(config.LOG_FILE_PATH) or ".", exist_ok=True)
        file_handler = logging.FileHandler(config.LOG_FILE_PATH)
        file_handler.setFormatter(logging.Formatter("%(asctime)s | %(levelname)s | %(message)s"))
        logger.addHandler(file_handler)
    except Exception:
        pass
    stream_handler = logging.StreamHandler()
    stream_handler.setFormatter(logging.Formatter("%(asctime)s | %(levelname)s | %(message)s"))
    logger.addHandler(stream_handler)
    return logger

def _save_fallback_snapshot(content: bytes, source_url: str):
    try:
        os.makedirs(os.path.dirname(config.FALLBACK_SNAPSHOT_PATH) or ".", exist_ok=True)
        with open(config.FALLBACK_SNAPSHOT_PATH, "wb") as f:
            f.write(content)
        meta = {"fetched_at": datetime.now().strftime("%d %b %Y, %H:%M:%S"), "source_url": source_url}
        with open(config.FALLBACK_META_PATH, "w") as f:
            json.dump(meta, f)
    except Exception as e:
        get_logger().warning(f"Could not persist fallback snapshot: {e}")

def _load_fallback_snapshot():
    try:
        if not (os.path.exists(config.FALLBACK_SNAPSHOT_PATH) and os.path.exists(config.FALLBACK_META_PATH)):
            return None
        with open(config.FALLBACK_SNAPSHOT_PATH, "rb") as f:
            content = f.read()
        with open(config.FALLBACK_META_PATH, "r") as f:
            meta = json.load(f)
        return content, meta
    except Exception as e:
        get_logger().warning(f"Could not load fallback snapshot: {e}")
        return None

def fetch_workbook_hardened(url: str, cache_key: str, _fetch_fn):
    logger = get_logger()
    try:
        content = _fetch_fn(url, cache_key)
        _save_fallback_snapshot(content, url)
        logger.info(f"Workbook fetched successfully from remote data stream: {url}")
        return content, {
            "source": "live", "fetched_at": datetime.now().strftime("%d %b %Y, %H:%M:%S"),
            "source_url": url, "warning": None,
        }
    except Exception as e:
        logger.error(f"Live pipeline connection drop for {url}: {e}")
        fallback = _load_fallback_snapshot()
        if fallback is not None:
            content, meta = fallback
            meta = dict(meta)
            meta["warning"] = f"Live feed offline ({e}); showing baseline backup file generated at {meta.get('fetched_at')}."
            meta["source"] = "fallback"
            logger.warning(meta["warning"])
            return content, meta
        logger.error("No fallback operational snapshot database found on disk storage systems.")
        raise

def validate_schema(file_bytes: bytes) -> list:
    issues = []
    try:
        wb = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as e:
        return [f"Could not verify file schema structural mappings: {e}"]

    sheet_names = wb.sheetnames
    for required in config.REQUIRED_SHEETS:
        if required not in sheet_names:
            issues.append(f"Required data layer workspace layout missing: '{required}'")

    label_col_map = {"H&S": 3, "Environment": 4}
    for sheet_name, required_labels in config.REQUIRED_KPI_LABELS.items():
        if sheet_name not in sheet_names:
            continue
        ws = wb[sheet_name]
        col_idx = label_col_map.get(sheet_name, 3)
        found_labels = set()
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx, values_only=True):
            val = row[0] if row else None
            if isinstance(val, str):
                found_labels.add(re.sub(r"\s+", " ", val).strip().lower())
        for label in required_labels:
            norm = re.sub(r"\s+", " ", label).strip().lower()
            if norm not in found_labels:
                issues.append(f"[{sheet_name}] Mandatory row schema target missing: '{label}'")
    wb.close()
    return issues

def resolve_latest_file_url(fallback_url: str) -> str:
    if not config.GITHUB_DATA_FOLDER:
        return fallback_url
    logger = get_logger()
    try:
        api_url = config.GITHUB_API_CONTENTS_URL.format(path=config.GITHUB_DATA_FOLDER.strip("/"))
        resp = requests.get(api_url, timeout=15)
        resp.raise_for_status()
        items = resp.json()
        files = [f for f in items if isinstance(f, dict) and f.get("name", "").lower().endswith(".xlsx")]
        latest_pattern = re.compile(config.LATEST_FILENAME_PATTERN, re.IGNORECASE)
        candidates = [f for f in files if latest_pattern.match(f["name"])]
        if not candidates:
            candidates = sorted(files, key=lambda f: f["name"])
        if not candidates:
            return fallback_url
        chosen = candidates[-1]
        return chosen.get("download_url") or fallback_url
    except Exception as e:
        logger.error(f"Auto-detect of folder file layouts fell back onto defaults: {e}")
        return fallback_url
