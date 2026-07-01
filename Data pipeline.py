"""
Phase 4 — Data Pipeline Hardening
===================================
This module wraps (does NOT replace) the existing Phase 1-3 ingestion
functions in app.py (`fetch_workbook_bytes`, `parse_workbook`). It adds:

1. Fallback caching   -> if a live GitHub fetch fails, serve the last
                          known-good workbook instead of hard-failing.
2. Schema validation  -> detect missing sheets / missing KPI rows before
                          they silently show up as "N/A" on the dashboard.
3. Multi-file support -> optionally auto-detect the active "latest*.xlsx"
                          file inside a GitHub folder, treating older
                          files in that folder as read-only archive/history.
4. Logging            -> every ingestion attempt (success, fallback, or
                          hard failure) is logged to disk + stderr.

None of the existing parsing/KPI-calculation logic in app.py is duplicated
here; this module only adds a safety layer around it.
"""

import os
import re
import json
import logging
from io import BytesIO
from datetime import datetime

import requests
from openpyxl import load_workbook

import config


# =============================================================================
# Logging
# =============================================================================
def get_logger() -> logging.Logger:
    logger = logging.getLogger("gnsc_dashboard")
    if logger.handlers:
        return logger
    logger.setLevel(getattr(logging, config.LOG_LEVEL, logging.INFO))
    try:
        os.makedirs(os.path.dirname(config.LOG_FILE_PATH) or ".", exist_ok=True)
        file_handler = logging.FileHandler(config.LOG_FILE_PATH)
        file_handler.setFormatter(logging.Formatter("%(asctime)s | %(levelname)s | %(message)s"))
        logger.addHandler(file_handler)
    except Exception:
        pass  # file logging is best-effort; never block the app on it
    stream_handler = logging.StreamHandler()
    stream_handler.setFormatter(logging.Formatter("%(asctime)s | %(levelname)s | %(message)s"))
    logger.addHandler(stream_handler)
    return logger


# =============================================================================
# Fallback snapshot persistence (Phase 4 §1)
# =============================================================================
def _save_fallback_snapshot(content: bytes, source_url: str):
    try:
        os.makedirs(os.path.dirname(config.FALLBACK_SNAPSHOT_PATH) or ".", exist_ok=True)
        with open(config.FALLBACK_SNAPSHOT_PATH, "wb") as f:
            f.write(content)
        meta = {"fetched_at": datetime.now().isoformat(), "source_url": source_url}
        with open(config.FALLBACK_META_PATH, "w") as f:
            json.dump(meta, f)
    except Exception as e:
        get_logger().warning(f"Could not persist fallback snapshot: {e}")


def _load_fallback_snapshot():
    try:
        if not (os.path.exists(config.FALLBACK_SNAPSHOT_PATH) and os.path.exists(config.FALLBACK_META_PATH)):
            return None
        with open(config.FALLBACK_SNAPSHOT_PATH, "rb") as f:
            content = f.read()
        with open(config.FALLBACK_META_PATH, "r") as f:
            meta = json.load(f)
        return content, meta
    except Exception as e:
        get_logger().warning(f"Could not load fallback snapshot: {e}")
        return None


def fetch_workbook_hardened(url: str, cache_key: str, _fetch_fn):
    """
    Fetches the workbook via the EXISTING cached fetch function
    (`_fetch_fn` — pass in app.py's `fetch_workbook_bytes`, unchanged) and
    wraps it with a persistent last-known-good fallback.

    Returns: (file_bytes, meta) where meta = {
        "source": "live" | "fallback",
        "fetched_at": iso timestamp,
        "source_url": str,
        "warning": str | None,
    }
    """
    logger = get_logger()
    try:
        content = _fetch_fn(url, cache_key)
        _save_fallback_snapshot(content, url)
        logger.info(f"Workbook fetched live from {url} ({len(content)} bytes).")
        return content, {
            "source": "live", "fetched_at": datetime.now().isoformat(),
            "source_url": url, "warning": None,
        }
    except Exception as e:
        logger.error(f"Live fetch failed for {url}: {e}")
        fallback = _load_fallback_snapshot()
        if fallback is not None:
            content, meta = fallback
            meta = dict(meta)
            meta["warning"] = f"Live fetch failed ({e}); showing last cached snapshot from {meta.get('fetched_at')}."
            meta["source"] = "fallback"
            logger.warning(meta["warning"])
            return content, meta
        logger.error("No fallback snapshot available; ingestion has no data to serve.")
        raise


# =============================================================================
# Schema / data validation (Phase 4 §1)
# =============================================================================
def validate_schema(file_bytes: bytes) -> list:
    """
    Lightweight, read-only validation pass over the workbook structure.
    Returns a list of human-readable issue strings (empty list = all good).
    Does not affect parse_workbook()'s own (separate, already-working) parse.
    """
    issues = []
    try:
        wb = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as e:
        return [f"Could not open workbook for validation: {e}"]

    sheet_names = wb.sheetnames
    for required in config.REQUIRED_SHEETS:
        if required not in sheet_names:
            issues.append(f"Missing required sheet: '{required}'")

    label_col_map = {"H&S": 3, "Environment": 4}  # 1-indexed: column C / D
    for sheet_name, required_labels in config.REQUIRED_KPI_LABELS.items():
        if sheet_name not in sheet_names:
            continue
        ws = wb[sheet_name]
        col_idx = label_col_map.get(sheet_name)
        found_labels = set()
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx, values_only=True):
            val = row[0] if row else None
            if isinstance(val, str):
                found_labels.add(re.sub(r"\s+", " ", val).strip().lower())
        for label in required_labels:
            norm = re.sub(r"\s+", " ", label).strip().lower()
            if norm not in found_labels:
                issues.append(f"[{sheet_name}] Missing expected KPI row: '{label}'")

    wb.close()
    if issues:
        get_logger().warning(f"Schema validation found {len(issues)} issue(s): {issues}")
    return issues


# =============================================================================
# Multi-file support: auto-detect latest workbook + archive listing (Phase 4 §2)
# =============================================================================
def list_github_files(folder: str) -> list:
    """Lists .xlsx files in a GitHub repo folder via the Contents API."""
    api_url = config.GITHUB_API_CONTENTS_URL.format(path=folder.strip("/"))
    resp = requests.get(api_url, timeout=15)
    resp.raise_for_status()
    items = resp.json()
    xlsx_files = [f for f in items if isinstance(f, dict) and f.get("name", "").lower().endswith(".xlsx")]
    get_logger().info(f"Found {len(xlsx_files)} .xlsx file(s) in '{folder or '/'}'.")
    return xlsx_files


def resolve_latest_file_url(fallback_url: str) -> str:
    """
    If `config.GITHUB_DATA_FOLDER` is set, auto-detects the active
    'latest*.xlsx' file inside that folder (older files are treated as
    read-only archive/history). If not configured, or if auto-detection
    fails for any reason, returns `fallback_url` unchanged — this keeps
    the current single-file deployment 100% backward compatible.
    """
    if not config.GITHUB_DATA_FOLDER:
        return fallback_url

    logger = get_logger()
    try:
        files = list_github_files(config.GITHUB_DATA_FOLDER)
        latest_pattern = re.compile(config.LATEST_FILENAME_PATTERN, re.IGNORECASE)
        candidates = [f for f in files if latest_pattern.match(f["name"])]
        if not candidates:
            # No file explicitly named "latest*" -- assume date-stamped
            # filenames and take the alphabetically-last one.
            candidates = sorted(files, key=lambda f: f["name"])
        if not candidates:
            logger.warning(f"No .xlsx files found in '{config.GITHUB_DATA_FOLDER}'; using fallback URL.")
            return fallback_url
        chosen = candidates[-1]
        raw_url = chosen.get("download_url") or fallback_url
        logger.info(f"Auto-detected active workbook: {chosen['name']}")
        return raw_url
    except Exception as e:
        logger.error(f"Auto-detect of latest file failed ({e}); using fallback URL.")
        return fallback_url


def list_archive_files() -> list:
    """
    Returns metadata for previous-month workbooks sitting in the archive
    sub-folder (`config.GITHUB_DATA_FOLDER/archive`). Read-only — does not
    affect the active dashboard's data path. Intended for future
    history/drill-down features.
    """
    if not config.GITHUB_DATA_FOLDER:
        return []
    archive_path = f"{config.GITHUB_DATA_FOLDER.strip('/')}/{config.ARCHIVE_FOLDER_NAME}"
    try:
        return list_github_files(archive_path)
    except Exception as e:
        get_logger().warning(f"Could not list archive folder '{archive_path}': {e}")
        return []
