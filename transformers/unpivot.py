"""
Metadata-Driven Layout Transformer Subsystem.
Restores both normalized and unpivoted structural dataframe parsers for absolute backward compatibility.
"""
import re
from io import BytesIO
import numpy as np
import pandas as pd
import openpyxl
import streamlit as st
import config

def _clean_string(val) -> str:
    if val is None or pd.isna(val):
        return ""
    return str(val).strip()

def _parse_numeric(val) -> float:
    if val is None or pd.isna(val) or str(val).strip() == "":
        return np.nan
    try:
        if isinstance(val, (int, float)):
            return float(val)
        return float(str(val).replace(",", "").strip())
    except (ValueError, TypeError):
        return np.nan

@st.cache_data(ttl=config.RAW_FILE_CACHE_TTL_SECONDS, show_spinner=False)
def melt_wide_sheet_to_long_cached(df_raw: pd.DataFrame) -> pd.DataFrame:
    """Legacy unpivot parsing routine preserved to fulfill historic application calls."""
    timeline_cols = []
    pattern = re.compile(r"(2025|2026|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec|Jan|Feb|Mar)", re.IGNORECASE)
    for col in df_raw.columns:
        col_str = str(col)
        if pattern.search(col_str):
            if "YTD" not in col_str and "Target" not in col_str:
                timeline_cols.append(col)

    label_col = None
    for col in df_raw.columns:
        sample = df_raw[col].dropna().astype(str).tolist()
        if any(any(k in s for k in ["Volume", "Fatalities", "waste", "Usage", "consumption", "withdrawal"]) for s in sample):
            label_col = col
            break
    if label_col is None:
        label_col = df_raw.columns[2] if len(df_raw.columns) > 2 else df_raw.columns[0]

    records = []
    raw_records = df_raw[[label_col] + timeline_cols].to_dict(orient="records")
    for row in raw_records:
        metric_raw = str(row[label_col]).strip()
        if not metric_raw or metric_raw.lower() in ["nan", "none", "environment monthly kpi"]:
            continue
        for period in timeline_cols:
            val = row[period]
            v_num = _parse_numeric(val)
            period_str = str(period).split(" ")[0] if "-" in str(period) else str(period)
            records.append({
                "Metric": metric_raw,
                "Month": period_str,
                "Value": v_num
            })
    return pd.DataFrame(records)

@st.cache_data(ttl=config.RAW_FILE_CACHE_TTL_SECONDS, show_spinner=False)
def infer_and_melt_workbook_metadata(file_bytes: bytes) -> dict:
    """Modern data-driven coordinate structure discovery parser."""
    wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
    all_records = []
    units_registry = {}
    catalog_tree = {}

    date_regex = re.compile(r"(\b\d{4}-\d{2}-\d{2}\b|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec|Jan|Feb|Mar)", re.IGNORECASE)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        max_row = ws.max_row
        max_column = ws.max_column
        
        if max_row < 4 or max_column < 5:
            continue

        row2 = [ws.cell(row=2, column=c).value for c in range(1, max_column + 1)]
        row3 = [ws.cell(row=3, column=c).value for c in range(1, max_column + 1)]

        for c in range(1, len(row2)):
            if row2[c] is None: row2[c] = row2[c-1]
        for c in range(1, len(row3)):
            if row3[c] is None: row3[c] = row3[c-1]

        timeline_mapping = {}
        target_col_idx = None
        ytd_col_idx = None
        mtd_col_idx = None

        for c in range(5, max_column + 1):
            val_r2 = _clean_string(row2[c-1])
            val_r3 = _clean_string(row3[c-1])
            
            if date_regex.search(val_r2) and "ytd" not in val_r2.lower() and "target" not in val_r2.lower():
                timeline_mapping[c] = val_r2
            elif date_regex.search(val_r3) and "ytd" not in val_r3.lower() and "target" not in val_r3.lower():
                timeline_mapping[c] = val_r3
                
            if "target" in val_r2.lower() or "target" in val_r3.lower():
                target_col_idx = c
            if "ytd" in val_r2.lower() or "ytd" in val_r3.lower():
                ytd_col_idx = c
            if "mtd" in val_r2.lower() or "mtd" in val_r3.lower():
                mtd_col_idx = c

        if not timeline_mapping:
            continue

        current_category = ""
        current_subcategory = ""

        for r in range(4, max_row + 1):
            cell_b = ws.cell(row=r, column=2).value
            cell_c = ws.cell(row=r, column=3).value
            cell_d = ws.cell(row=r, column=4).value

            if cell_b is not None and _clean_string(cell_b) != "": 
                current_category = str(cell_b).strip()
            if cell_c is not None and _clean_string(cell_c) != "": 
                current_subcategory = str(cell_c).strip()
                
            if cell_d is None or _clean_string(cell_d).lower() in ["nan", "none", ""]:
                continue
                
            metric_name = str(cell_d).strip()

            inferred_unit = ""
            for c_idx in timeline_mapping.keys():
                unit_candidate = _clean_string(row3[c_idx-1])
                if unit_candidate and not date_regex.search(unit_candidate) and unit_candidate.lower() not in ["ytd", "target", "mtd", "bu"]:
                    inferred_unit = unit_candidate
                    break
            
            if not inferred_unit and target_col_idx:
                inferred_unit = _clean_string(row3[target_col_idx-1])
                
            if not inferred_unit or inferred_unit.lower() in ["ytd", "target", "mtd"]:
                inferred_unit = "Units"

            units_registry[metric_name] = inferred_unit

            target_val = _parse_numeric(ws.cell(row=r, column=target_col_idx).value) if target_col_idx else np.nan
            ytd_val = _parse_numeric(ws.cell(row=r, column=ytd_col_idx).value) if ytd_col_idx else np.nan
            mtd_val = _parse_numeric(ws.cell(row=r, column=mtd_col_idx).value) if mtd_col_idx else np.nan

            for col_idx, period_label in timeline_mapping.items():
                numeric_val = _parse_numeric(ws.cell(row=r, column=col_idx).value)
                
                inferred_year = config.CURRENT_FISCAL_YEAR
                year_match = re.search(r"\d{4}", period_label)
                if year_match:
                    inferred_year = year_match.group(0)

                if "-" in period_label:
                    try:
                        period_clean = datetime.strptime(period_label.split(" ")[0], "%Y-%m-%d").strftime("%b-%y")
                    except Exception:
                        period_clean = period_label
                else:
                    period_clean = period_label

                all_records.append({
                    "Sheet": str(sheet_name).strip(),
                    "Category": current_category if current_category else str(sheet_name).strip(),
                    "Subcategory": current_subcategory if current_subcategory else "General Operations",
                    "Metric": metric_name,
                    "Month": period_clean,
                    "Value": numeric_val,
                    "Unit": inferred_unit,
                    "Target": target_val,
                    "YTD": ytd_val,
                    "MTD": mtd_val,
                    "Year": inferred_year
                })

    df_long = pd.DataFrame(all_records)
    
    if not df_long.empty:
        distinct_categories = df_long["Category"].unique().tolist()
        for cat in distinct_categories:
            catalog_tree[cat] = sorted(df_long[df_long["Category"] == cat]["Metric"].unique().tolist())
        sorted_periods = sorted(df_long["Month"].unique().tolist(), key=lambda x: ("2026" in str(x), x))
    else:
        sorted_periods = []

    return {
        "long_df": df_long,
        "catalog": catalog_tree,
        "periods": sorted_periods,
        "units": units_registry
    }
